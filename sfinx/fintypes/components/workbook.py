from io import BytesIO

from openpyxl import load_workbook

from sfinx.fintypes.components.sheet import FinSheet


class FinWorkbook:
    """
    Represents an MS Excel workbook including one or more worksheets.
    """

    def __init__(self, fl, from_path=True):
        f = fl if from_path else BytesIO(fl)
        self.wb = load_workbook(f, data_only=True)
        self.sheets = {}
        self._load_sheets()

    def _load_sheets(self):
        """
        Loads individual worksheets into FinSheet objects.
        Keeps track of them via the self.sheets map
        """
        for sn in self.wb.sheetnames:
            st = self.wb[sn]
            if st.max_row > 1 and st.max_column > 1:
                self.sheets[sn] = FinSheet(sn, st)
