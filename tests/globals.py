import os

from openpyxl import load_workbook

from sfinx.fintypes.components.cell import FinTabCell
from sfinx.fintypes.components.workbook import FinWorkbook
from sfinx.processors.metric_normalizer import FinTabMetricNormalizer
from sfinx.processors.period_normalizer import FinTabPeriodNormalizer

sample_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "samples", "sample1.xlsx")
sample_workbook = load_workbook(sample_path)
sample_toc = sample_workbook["Table of Contents (Hyperlinks)"]
sample_table = sample_workbook["OF segment accounts"]
sample_wb = FinWorkbook(sample_path, from_path=True)
sample_fin_table = sample_wb.sheets["OF segment accounts"]
sample_cell, footnote_flags = FinTabCell.clean_footnotes(sample_table["D4"])
sample_cell_i = 3
sample_cell_j = 3
sample_period_norm = FinTabPeriodNormalizer(sample_wb)
sample_metric_norm = FinTabMetricNormalizer(sample_wb)
